"""
Экспорт в Excel
"""

import math
import pandas as pd
from io import BytesIO
from typing import Optional


def safe_float(value, default=0) -> float:
    """Безопасное преобразование в float: NaN, None, строки → default"""
    if value is None:
        return default
    if isinstance(value, (int, float)):
        if math.isnan(value) or math.isinf(value):
            return default
        return float(value)
    try:
        return float(value)
    except (ValueError, TypeError):
        return default


def safe_str(value) -> str:
    """Безопасное преобразование в строку: NaN, None → пустая строка"""
    if value is None:
        return ''
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return ''
    s = str(value)
    if s in ('nan', 'None', '<NA>', 'NaT'):
        return ''
    return s

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def export_kp_to_excel(df: pd.DataFrame, contract_type: str = "КП") -> bytes:
    """
    Экспорт КП в Excel

    Args:
        df: DataFrame с данными КП
        contract_type: Тип контракта (РБ/ФБ)

    Returns:
        Байты Excel файла
    """
    output = BytesIO()

    if not HAS_OPENPYXL:
        # Fallback на pandas
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='КП')
        output.seek(0)
        return output.getvalue()

    wb = Workbook()
    ws = wb.active
    ws.title = "Коммерческое предложение"

    # Стили
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    number_format = '#,##0.00'
    money_format = '#,##0.00 ₽'

    # Заголовок документа
    ws.merge_cells('A1:G1')
    ws['A1'] = f"КОММЕРЧЕСКОЕ ПРЕДЛОЖЕНИЕ ({contract_type})"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Колонки для КП
    columns = ['№', 'Наименование', 'Описание', 'Ед.изм.', 'Кол-во', 'Цена за ед.', 'Сумма']

    # Заголовки таблицы
    start_row = 3
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=col_name)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Данные
    row_idx = start_row + 1
    total_sum = 0

    for _, row in df.iterrows():
        ws.cell(row=row_idx, column=1, value=row.get('№', row_idx - start_row)).border = border

        name_cell = ws.cell(row=row_idx, column=2, value=safe_str(row.get('Наименование', '')))
        name_cell.border = border
        name_cell.alignment = Alignment(wrap_text=True, vertical='top')

        desc_cell = ws.cell(row=row_idx, column=3, value=safe_str(row.get('Описание', '')))
        desc_cell.border = border
        desc_cell.alignment = Alignment(wrap_text=True, vertical='top')

        ws.cell(row=row_idx, column=4, value=safe_str(row.get('Ед.изм.', 'кг'))).border = border

        qty = safe_float(row.get('Кол-во', 0))
        qty_cell = ws.cell(row=row_idx, column=5, value=qty)
        qty_cell.border = border
        qty_cell.number_format = '#,##0'

        price = safe_float(row.get('Наша цена', 0))
        price_cell = ws.cell(row=row_idx, column=6, value=price)
        price_cell.border = border
        price_cell.number_format = number_format

        sum_value = round(price * qty, 2)
        sum_cell = ws.cell(row=row_idx, column=7, value=sum_value)
        sum_cell.border = border
        sum_cell.number_format = money_format

        total_sum += sum_value
        row_idx += 1

    # Итого
    ws.merge_cells(f'A{row_idx}:F{row_idx}')
    ws.cell(row=row_idx, column=1, value="ИТОГО:").font = Font(bold=True)
    ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal='right')
    ws.cell(row=row_idx, column=1).border = border

    total_cell = ws.cell(row=row_idx, column=7, value=total_sum)
    total_cell.font = Font(bold=True)
    total_cell.border = border
    total_cell.number_format = money_format

    # Ширина колонок
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 18

    wb.save(output)
    output.seek(0)
    return output.getvalue()


def export_economics_to_excel(df: pd.DataFrame, contract_type: str = "КП") -> bytes:
    """
    Экспорт расчёта экономики в Excel

    Args:
        df: DataFrame с данными КП
        contract_type: Тип контракта (РБ/ФБ)

    Returns:
        Байты Excel файла
    """
    output = BytesIO()

    if not HAS_OPENPYXL:
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Экономика')
        output.seek(0)
        return output.getvalue()

    wb = Workbook()
    ws = wb.active
    ws.title = "Расчёт экономики"

    # Стили
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    number_format = '#,##0.00'
    percent_format = '0.00%'
    money_format = '#,##0.00 ₽'

    # Заголовок
    ws.merge_cells('A1:L1')
    ws['A1'] = f"РАСЧЁТ ЭКОНОМИКИ ПОСТАВКИ ({contract_type})"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Колонки
    columns = [
        '№', 'Наименование', 'Ед.изм.', 'Кол-во',
        'Себестоимость', 'Наша цена', 'Цена конкурента',
        'Сумма', 'Маржа', 'Маржа %', 'Прибыль'
    ]

    # Заголовки
    start_row = 3
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=col_name)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Данные
    row_idx = start_row + 1
    totals = {'Сумма': 0, 'Прибыль': 0}

    for _, row in df.iterrows():
        ws.cell(row=row_idx, column=1, value=row.get('№', '')).border = border
        ws.cell(row=row_idx, column=2, value=safe_str(row.get('Наименование', ''))).border = border
        ws.cell(row=row_idx, column=3, value=safe_str(row.get('Ед.изм.', 'кг'))).border = border

        qty = safe_float(row.get('Кол-во', 0))
        qty_cell = ws.cell(row=row_idx, column=4, value=qty)
        qty_cell.border = border
        qty_cell.number_format = '#,##0'

        cost = safe_float(row.get('Себестоимость', 0))
        cost_cell = ws.cell(row=row_idx, column=5, value=cost)
        cost_cell.border = border
        cost_cell.number_format = number_format

        our_price = safe_float(row.get('Наша цена', 0))
        our_cell = ws.cell(row=row_idx, column=6, value=our_price)
        our_cell.border = border
        our_cell.number_format = number_format

        comp_price = safe_float(row.get('Цена конкурента', 0))
        comp_cell = ws.cell(row=row_idx, column=7, value=comp_price)
        comp_cell.border = border
        comp_cell.number_format = number_format

        sum_value = round(our_price * qty, 2)
        sum_cell = ws.cell(row=row_idx, column=8, value=sum_value)
        sum_cell.border = border
        sum_cell.number_format = money_format
        totals['Сумма'] += sum_value

        margin = our_price - cost
        margin_cell = ws.cell(row=row_idx, column=9, value=margin)
        margin_cell.border = border
        margin_cell.number_format = number_format

        margin_pct = (margin / our_price) if our_price > 0 else 0
        margin_pct_cell = ws.cell(row=row_idx, column=10, value=margin_pct)
        margin_pct_cell.border = border
        margin_pct_cell.number_format = percent_format

        profit = margin * qty
        profit_cell = ws.cell(row=row_idx, column=11, value=profit)
        profit_cell.border = border
        profit_cell.number_format = money_format
        totals['Прибыль'] += profit

        row_idx += 1

    # Итого
    ws.merge_cells(f'A{row_idx}:G{row_idx}')
    ws.cell(row=row_idx, column=1, value="ИТОГО:").font = Font(bold=True)
    ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal='right')
    ws.cell(row=row_idx, column=1).border = border

    total_sum_cell = ws.cell(row=row_idx, column=8, value=totals['Сумма'])
    total_sum_cell.font = Font(bold=True)
    total_sum_cell.border = border
    total_sum_cell.number_format = money_format

    total_profit_cell = ws.cell(row=row_idx, column=11, value=totals['Прибыль'])
    total_profit_cell.font = Font(bold=True)
    total_profit_cell.border = border
    total_profit_cell.number_format = money_format

    # Ширина колонок
    widths = [5, 35, 8, 10, 14, 14, 14, 16, 12, 10, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    wb.save(output)
    output.seek(0)
    return output.getvalue()
